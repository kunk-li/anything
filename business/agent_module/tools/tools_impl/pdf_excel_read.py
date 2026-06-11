# -*- coding: utf-8 -*-
"""
Tools: pdf_read + excel_read (Task TTTT-4 #141)

读取本地 PDF/Excel 文件并返文本/单元格数据.
依赖: pypdf, openpyxl (在 requirements.txt 里).

安全:
  - 文件路径限制在 uploads/ + source_docs/ + /tmp 内, 防穿越
  - 文本截断 max_chars 防爆 prompt
"""
from __future__ import annotations

import os
from pathlib import Path
from typing import Any, Dict, List, Optional


# 允许读取的目录前缀 (按 cwd 解析). path 必须在这些里面.
_ALLOWED_PREFIXES = ("uploads", "source_docs", "documents", "run/uploads")

# 扫描版 PDF (无文字层) 自动渲染参数: 页数上限防止几百页扫描件逐页跑多模态
SCAN_RENDER_MAX_PAGES = 5
SCAN_RENDER_ZOOM = 2.0  # ~144 DPI, OCR 可读性与图片体积的折中


def _resolve_safe_path(file_path: str) -> Optional[Path]:
    """归一化并校验路径; 不允许 .. 穿越或绝对路径出沙盒."""
    if not file_path or not isinstance(file_path, str):
        return None
    p = Path(file_path)
    if p.is_absolute():
        # 绝对路径: 必须在 cwd 内
        try:
            rp = p.resolve(strict=False)
        except Exception:
            return None
        try:
            rp.relative_to(Path.cwd().resolve())
        except ValueError:
            return None
        return rp if rp.exists() else None
    # 相对路径: 拼到 cwd, 校验前缀
    candidate = (Path.cwd() / p).resolve()
    try:
        candidate.relative_to(Path.cwd().resolve())
    except ValueError:
        return None
    # 至少匹配一个允许的前缀 (再宽就太开放)
    rel = candidate.relative_to(Path.cwd().resolve())
    if not any(str(rel).replace("\\", "/").startswith(pre) for pre in _ALLOWED_PREFIXES):
        # 给点提示而不是直接拒, 让 LLM 知道路径不对
        return None
    return candidate if candidate.exists() else None


def pdf_read(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Input: {"file_path": str, "max_chars": int = 6000, "page_range": "1-5" / "all"}
    Output: {data: {text, num_pages, file_path, truncated}}
    """
    trace_id = payload.get("trace_id")
    file_path = payload.get("file_path", "")
    max_chars = int(payload.get("max_chars") or 6000)
    page_range = payload.get("page_range") or "all"

    safe = _resolve_safe_path(file_path)
    if safe is None:
        return _err(
            "INVALID_PATH",
            f"file_path 无效或不在允许范围 (uploads/ source_docs/ documents/). 收到: {file_path!r}",
            trace_id,
        )

    try:
        from pypdf import PdfReader
    except ImportError:
        return _err("MISSING_DEPS", "pypdf 未安装, 跑 pip install pypdf", trace_id)

    try:
        reader = PdfReader(str(safe))
        num_pages = len(reader.pages)
    except Exception as e:
        return _err("READ_FAILED", f"打开 PDF 失败: {e}", trace_id)

    # 解析 page_range
    page_idxs: List[int] = []
    pr = page_range.strip().lower()
    if pr in ("all", ""):
        page_idxs = list(range(num_pages))
    else:
        try:
            if "-" in pr:
                a, b = pr.split("-", 1)
                start = max(1, int(a)) - 1
                end = min(num_pages, int(b))
                page_idxs = list(range(start, end))
            else:
                page_idxs = [int(pr) - 1]
        except Exception:
            return _err("PARAM_INVALID", f"page_range 格式错: {page_range!r} (用 'all' 或 '1-5')", trace_id)

    parts: List[str] = []
    total = 0
    truncated = False
    for i in page_idxs:
        if i < 0 or i >= num_pages:
            continue
        try:
            t = reader.pages[i].extract_text() or ""
        except Exception:
            t = ""
        if t:
            chunk = f"\n--- Page {i+1} ---\n{t}"
            if total + len(chunk) > max_chars:
                parts.append(chunk[: max_chars - total])
                truncated = True
                break
            parts.append(chunk)
            total += len(chunk)
        if total >= max_chars:
            truncated = True
            break

    text = "".join(parts).strip()

    # 扫描版兜底: 没有文字层 (纯图像 PDF) → 渲染页面为图片, 指示 Agent 转
    # image_describe 多模态识别。文字层哪怕只有几个字也按正常文本走。
    if not text:
        page_images, render_note = _render_scanned_pages(safe, page_idxs)
        if page_images:
            return {
                "code": "SUCCESS", "message": "ok",
                "data": {
                    "text": "",
                    "num_pages": num_pages,
                    "file_path": str(safe.relative_to(Path.cwd())),
                    "pages_read": len(page_idxs),
                    "scanned": True,
                    "page_images": page_images,
                    "description": (
                        f"PDF {safe.name} 是扫描版 (无文字层, 共 {num_pages} 页)。"
                        f"已渲染 {len(page_images)} 页为图片: {page_images}。"
                        f"请对每个图片路径调用 image_describe(image_path) 识别页面内容, "
                        f"再回答用户问题。" + (f" ({render_note})" if render_note else "")
                    ),
                },
                "trace_id": trace_id,
                "retryable": False,
                "details": None,
            }
        # 渲染不可用 (缺 pymupdf 等) → 如实说明, 不让 Agent 误以为是空白文档
        return {
            "code": "SUCCESS", "message": "ok",
            "data": {
                "text": "",
                "num_pages": num_pages,
                "file_path": str(safe.relative_to(Path.cwd())),
                "pages_read": len(page_idxs),
                "scanned": True,
                "page_images": [],
                "description": (
                    f"PDF {safe.name} 是扫描版 (无文字层, 共 {num_pages} 页), "
                    f"且页面渲染不可用 ({render_note}), 无法转多模态识别。"
                ),
            },
            "trace_id": trace_id,
            "retryable": False,
            "details": None,
        }

    return {
        "code": "SUCCESS", "message": "ok",
        "data": {
            "text": text,
            "num_pages": num_pages,
            "file_path": str(safe.relative_to(Path.cwd())),
            "pages_read": len(page_idxs),
            "truncated": truncated,
            "description": (
                f"读取 PDF {safe.name}: {num_pages} 页, "
                f"提取 {len(text)} 字符" + (" (已截断)" if truncated else "")
            ),
        },
        "trace_id": trace_id,
        "retryable": False,
        "details": None,
    }


def _render_scanned_pages(pdf_path: Path, page_idxs: List[int]) -> "tuple[List[str], str]":
    """扫描版 PDF → 页面 PNG。渲染到 PDF 同级 screenshots/ 子目录
    (upload-janitor 把它当工具产物按保留期回收, 不会永久堆积)。

    返回 (相对 cwd 的图片路径列表, 备注)。渲染库缺失/失败 → ([], 原因)。
    """
    try:
        import fitz  # PyMuPDF
    except ImportError:
        return [], "pymupdf 未安装, 跑 pip install pymupdf"

    todo = page_idxs[:SCAN_RENDER_MAX_PAGES]
    note = ""
    if len(page_idxs) > len(todo):
        note = (f"仅渲染前 {len(todo)} 页, 其余 {len(page_idxs) - len(todo)} 页"
                f'用 page_range 参数 (如 "6-10") 再取')

    out_dir = pdf_path.parent / "screenshots"
    images: List[str] = []
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
        with fitz.open(str(pdf_path)) as doc:
            mat = fitz.Matrix(SCAN_RENDER_ZOOM, SCAN_RENDER_ZOOM)
            for i in todo:
                if i < 0 or i >= doc.page_count:
                    continue
                pix = doc[i].get_pixmap(matrix=mat)
                out = out_dir / f"{pdf_path.stem}.scan.p{i + 1}.png"
                pix.save(str(out))
                try:
                    images.append(str(out.relative_to(Path.cwd())))
                except ValueError:
                    images.append(str(out))
    except Exception as e:
        return [], f"页面渲染失败: {e}"
    return images, note


def excel_read(payload: Dict[str, Any]) -> Dict[str, Any]:
    """
    Input:
        file_path: str
        sheet_name: str = ""    空 = 第一个 sheet
        max_rows: int = 100
        max_cols: int = 30
    Output: {data: {sheet_name, headers, rows, total_rows_in_sheet, file_path}}
    """
    trace_id = payload.get("trace_id")
    file_path = payload.get("file_path", "")
    sheet_name = (payload.get("sheet_name") or "").strip()
    max_rows = int(payload.get("max_rows") or 100)
    max_cols = int(payload.get("max_cols") or 30)

    safe = _resolve_safe_path(file_path)
    if safe is None:
        return _err(
            "INVALID_PATH",
            f"file_path 无效或不在允许范围. 收到: {file_path!r}",
            trace_id,
        )

    try:
        import openpyxl
    except ImportError:
        return _err("MISSING_DEPS", "openpyxl 未安装, 跑 pip install openpyxl", trace_id)

    try:
        wb = openpyxl.load_workbook(str(safe), read_only=True, data_only=True)
    except Exception as e:
        return _err("READ_FAILED", f"打开 Excel 失败: {e}", trace_id)

    try:
        sheet_names = wb.sheetnames
        if sheet_name:
            if sheet_name not in sheet_names:
                return _err(
                    "PARAM_INVALID",
                    f"sheet_name {sheet_name!r} 不存在; 可选: {sheet_names}",
                    trace_id,
                )
            ws = wb[sheet_name]
        else:
            ws = wb[sheet_names[0]]
            sheet_name = ws.title

        # 第 1 行当 header
        headers: List[str] = []
        rows: List[List[Any]] = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if r_idx == 0:
                headers = [str(c) if c is not None else "" for c in list(row)[:max_cols]]
                continue
            if r_idx > max_rows:
                break
            rows.append([
                (c if c is None or isinstance(c, (int, float, str, bool)) else str(c))
                for c in list(row)[:max_cols]
            ])

        return {
            "code": "SUCCESS", "message": "ok",
            "data": {
                "sheet_name": sheet_name,
                "all_sheets": sheet_names,
                "headers": headers,
                "rows": rows,
                "rows_read": len(rows),
                "file_path": str(safe.relative_to(Path.cwd())),
                "description": (
                    f"读取 Excel {safe.name} sheet {sheet_name!r}: "
                    f"{len(headers)} 列, {len(rows)} 行 (前 {max_rows} 行)"
                ),
            },
            "trace_id": trace_id,
            "retryable": False,
            "details": None,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _err(code: str, message: str, trace_id: Optional[str]) -> Dict[str, Any]:
    return {
        "code": code, "message": message, "data": None,
        "trace_id": trace_id, "retryable": False, "details": None,
    }


PDF_READ_DESCRIPTION = (
    '读取本地 PDF 文件抽取文本. input: {"file_path": str, '
    '"max_chars": int? (默 6000), "page_range": "all"|"1-5"|"3"?}. '
    "路径只允许在 uploads/ source_docs/ documents/ 下. "
    "返回 data.text + data.num_pages. 适合: PDF 总结/问答. "
    "扫描版 (无文字层) 会自动渲染页面图并返回 data.page_images — "
    "此时对每个路径调用 image_describe(image_path) 识别内容."
)

EXCEL_READ_DESCRIPTION = (
    '读取本地 Excel (.xlsx) 文件返回表头 + 数据行. input: {"file_path": str, '
    '"sheet_name": str? (默第一个), "max_rows": int? (默 100), '
    '"max_cols": int? (默 30)}. 路径只允许在 uploads/ source_docs/ documents/ 下. '
    "返回 data.headers + data.rows + data.all_sheets. 适合: 表格数据分析."
)
